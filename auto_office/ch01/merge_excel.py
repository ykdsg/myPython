from pathlib import Path, PurePath

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

src_path = './调查问卷'
dst_file = './result/结果.xlsx'

# 取得该目录下所有的xlsx格式文件
p = Path(src_path)
files = [x for x in p.iterdir() if PurePath(x).match('*.xlsx')]

# 准备一个列表存放读取结果
content = []

for file in files:
    # 用文件名作为每个用户的标识
    username = file.stem
    data = load_workbook(file)
    sheet = data.active

    # 取得每一项的结果
    answer1 = sheet.cell(row=5, column=5).value
    answer2 = sheet.cell(row=11, column=5).value
    temp = f'{username},{answer1},{answer2}'
    content.append(temp.split(','))
    print(temp)

# 准备写入文件的表头
table_header = ['员工姓名', '第一题', '第二题']

workbook = Workbook()
xlsheet = workbook.active
xlsheet.title = '统计结果'

# 写入表头
xlsheet.append(table_header)
# 取出每一行内容
for line in content:
    xlsheet.append(line)

# 保存最终结果
workbook.save(dst_file)
