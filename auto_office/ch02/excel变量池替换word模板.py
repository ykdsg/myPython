from docx import Document
from pathlib import Path, PurePath
import datetime

from openpyxl.reader.excel import load_workbook

today = datetime.date.today().strftime('%Y-%m-%d')

# 客户信息文件
customer = './邀请函样例文件/客户信息.xlsx'

# 邀请函模版
invitation = './邀请函样例文件/邀请函模版.docx'

# 邀请函路径
invitation_path = './邀请函样例文件/'

# 替换内容
replace_content = {
    '<姓名>': 'no_name',
    '<性别>': 'm_f',
    '<今天日期>': today
}


def generat_invitation():
    '''
    生成邀请函文件
    '''
    doc = Document(invitation)
    # 取出每一段
    for para in doc.paragraphs:
        for key, value in replace_content.items():
            if key in para.text:
                # 逐个关键字进行替换
                para.text = para.text.replace(key, value)

    file_name = PurePath(invitation_path, f"{replace_content['<姓名>']}.docx")

    doc.save(file_name)


def get_customer(customer_file: Path):
    '''
    获取邀请函信息
    '''
    # 从第一个sheet中取出客户信息
    data = load_workbook(customer_file)
    table = data.active

    # 取得客户数量
    customer_number = table.max_row

    for line in range(1, customer_number + 1):
        content = [cell.value for cell in table[line]]
        replace_content['<姓名>'] = content[0]
        replace_content['<性别>'] = content[1]
        print(replace_content)
        # {'<姓名>': '韩梅梅', '<性别>': '女士', '<今天日期>': '2021-01-01'}
        # {'<姓名>': '李雷', '<性别>': '先生', '<今天日期>': '2021-01-01'}
        generat_invitation()


if __name__ == '__main__':
    get_customer(customer)
